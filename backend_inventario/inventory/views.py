import logging
import json
from collections import defaultdict
from datetime import datetime
from io import BytesIO
from pathlib import Path
import pandas as pd
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.db.models import Q
from .models import ImportBatch, Product, InventoryRecord
from .services.analytics_service import (
    get_inventory_at_date_data,
    get_monthly_cuts_data,
    get_monthly_product_cuts_data,
    get_monthly_movements_data,
    get_product_analysis_data,
)
from .services.import_service import procesar_importacion_inventario
from .services.summary_service import get_inventory_summary_data

# ReportLab imports
from reportlab.lib import colors as rl_colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    BaseDocTemplate,
    Frame,
    HRFlowable,
    Image as RLImage,
    PageTemplate,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie as RLPie
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.legends import Legend
from reportlab.graphics.renderPDF import GraphicsFlowable

# openpyxl imports
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint as XLDataPoint

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Paleta corporativa GeoFlora — Rosa #EF7C91 · Blanco #FFFFFD · Cyan #4BC0D9
# ---------------------------------------------------------------------------
_CORP_HEADER_BG   = rl_colors.HexColor('#EF7C91')   # Rosa primario
_CORP_HEADER_BG2  = rl_colors.HexColor('#4BC0D9')   # Cyan secundario
_CORP_ROW_ALT     = rl_colors.HexColor('#FFF0F3')   # Rosa muy suave
_CORP_TOTAL_BG    = rl_colors.HexColor('#EBF9FC')   # Cyan muy suave
_CORP_TEXT        = rl_colors.HexColor('#1A1F2E')
_CORP_BORDER      = rl_colors.HexColor('#E5E7EB')
_CORP_WHITE       = rl_colors.white

_XL_HEADER_BG  = 'EF7C91'   # Rosa primario
_XL_HEADER_FG  = 'FFFFFF'
_XL_ROW_ALT    = 'FFF0F3'   # Rosa muy suave
_XL_TOTAL_BG   = 'EBF9FC'   # Cyan muy suave
_XL_BORDER_CLR = 'BDBDBD'
_XL_TEXT       = '1A1F2E'
_XL_ACCENT     = '4BC0D9'   # Cyan secundario

# ---------------------------------------------------------------------------
# Helpers de reportes
# ---------------------------------------------------------------------------

def _get_logo_path():
    """Localiza logo_geoflora.png relativo a este módulo. Retorna Path o None."""
    module_dir = Path(__file__).resolve().parent          # .../inventory/
    repo_root  = module_dir.parent.parent                 # geoo_controller/
    logo = repo_root / 'geo_inventario' / 'statics' / 'images' / 'logo_geoflora.png'
    if logo.exists():
        return logo
    local = module_dir / 'static' / 'logo_geoflora.png'
    if local.exists():
        return local
    logger.warning(f"Logo no encontrado en {logo}")
    return None


def _build_pdf_header_elements(logo_path, title, inventory_name, subtitle=None):
    """Genera flowables del encabezado corporativo: [Logo | Título + Inventario + Fecha] + separador."""
    rl_styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'HdrTitle', parent=rl_styles['Heading1'],
        fontName='Times-Bold', fontSize=15,
        textColor=_CORP_HEADER_BG2, spaceAfter=2,
    )
    inv_style = ParagraphStyle(
        'HdrInv', parent=rl_styles['Normal'],
        fontName='Times-Roman', fontSize=9,
        textColor=_CORP_TEXT,
    )
    date_style = ParagraphStyle(
        'HdrDate', parent=rl_styles['Normal'],
        fontName='Times-Italic', fontSize=8,
        textColor=rl_colors.HexColor('#3D4459'),
    )
    right_col = [
        Paragraph(title, title_style),
        Paragraph(f'Inventario: {inventory_name}', inv_style),
        Paragraph(f'Generado: {datetime.now().strftime("%Y-%m-%d %H:%M")}', date_style),
    ]
    if subtitle:
        right_col.append(Paragraph(subtitle, date_style))

    if logo_path:
        left_cell = RLImage(str(logo_path), width=3 * cm, height=1.5 * cm)
    else:
        left_cell = Paragraph('<b>GeoFlora</b>', title_style)

    hdr_table = Table([[left_cell, right_col]], colWidths=[3.8 * cm, None])
    hdr_table.setStyle(TableStyle([
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN',         (0, 0), (0, 0),   'CENTER'),
        ('ALIGN',         (1, 0), (1, 0),   'LEFT'),
        ('TOPPADDING',    (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    separator = HRFlowable(width='100%', thickness=2, color=_CORP_HEADER_BG, spaceAfter=8)
    return [hdr_table, Spacer(1, 4), separator]


def _create_pdf_doc(buffer, pagesize, report_title):
    """Crea BaseDocTemplate con footer 'Geo Inventario | Título | Página X | Fecha'."""
    w, h = pagesize
    margin = 1.5 * cm

    def _footer(canvas, doc):
        canvas.saveState()
        canvas.setFont('Times-Roman', 7)
        canvas.setFillColor(rl_colors.HexColor('#3D4459'))
        text = f'Geo Inventario  |  {report_title}  |  Página {doc.page}  |  {datetime.now().strftime("%Y-%m-%d")}'
        canvas.drawString(margin, 0.6 * cm, text)
        canvas.setStrokeColor(_CORP_HEADER_BG)
        canvas.setLineWidth(0.5)
        canvas.line(margin, 0.85 * cm, w - margin, 0.85 * cm)
        canvas.restoreState()

    frame = Frame(margin, 1.4 * cm, w - 2 * margin, h - 2 * margin, id='main')
    template = PageTemplate(id='main', frames=[frame], onPage=_footer)
    doc = BaseDocTemplate(
        buffer, pagesize=pagesize, pageTemplates=[template],
        leftMargin=margin, rightMargin=margin,
        topMargin=margin, bottomMargin=1.4 * cm,
    )
    return doc


def _apply_excel_header(ws, title, inventory_name, logo_path, header_rows=4):
    """Inserta logo, título y metadatos en las primeras filas del worksheet. Retorna fila de inicio de datos."""
    if logo_path:
        try:
            img = XLImage(str(logo_path))
            img.width = 120
            img.height = 60
            ws.add_image(img, 'A1')
        except Exception as e:
            logger.warning(f"No se pudo insertar logo en Excel: {e}")
    ws.row_dimensions[1].height = 48
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 15
    ws.row_dimensions[4].height = 10

    title_cell = ws['B1']
    title_cell.value = title
    title_cell.font = Font(name='Calibri', size=15, bold=True, color=_XL_ACCENT)
    title_cell.alignment = Alignment(vertical='center')

    ws['B2'].value = f'Inventario: {inventory_name}'
    ws['B2'].font = Font(name='Calibri', size=10, color=_XL_TEXT)

    ws['B3'].value = f'Generado: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws['B3'].font = Font(name='Calibri', size=9, italic=True, color='3D4459')

    return header_rows + 1  # primera fila disponible para datos


def _style_excel_table(ws, header_row, data_start, data_end, col_count):
    """Aplica estilos corporativos: header verde, filas alternadas, bordes."""
    header_fill = PatternFill(start_color=_XL_HEADER_BG, end_color=_XL_HEADER_BG, fill_type='solid')
    header_font = Font(color=_XL_HEADER_FG, bold=True, name='Calibri', size=11)
    alt_fill    = PatternFill(start_color=_XL_ROW_ALT, end_color=_XL_ROW_ALT, fill_type='solid')
    thin        = Side(border_style='thin', color=_XL_BORDER_CLR)
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center      = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col in range(1, col_count + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center
        cell.border    = cell_border

    for row_idx in range(data_start, data_end + 1):
        is_alt = (row_idx - data_start) % 2 == 1
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_idx, column=col)
            if is_alt:
                cell.fill = alt_fill
            cell.border    = cell_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)


def _pdf_table_style(data_len):
    """Retorna TableStyle corporativo con zebra striping para tablas PDF."""
    style_cmds = [
        ('BACKGROUND',    (0, 0), (-1, 0),  _CORP_HEADER_BG),
        ('TEXTCOLOR',     (0, 0), (-1, 0),  _CORP_WHITE),
        ('FONTNAME',      (0, 0), (-1, 0),  'Times-Bold'),
        ('FONTSIZE',      (0, 0), (-1, 0),  9),
        ('TOPPADDING',    (0, 0), (-1, 0),  5),
        ('BOTTOMPADDING', (0, 0), (-1, 0),  5),
        ('LINEBELOW',     (0, 0), (-1, 0),  1.0, _CORP_HEADER_BG2),
        ('BACKGROUND',    (0, 1), (-1, -1), _CORP_WHITE),
        ('GRID',          (0, 0), (-1, -1), 0.4, _CORP_BORDER),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING',   (0, 0), (-1, -1), 4),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
    ]
    for i in range(1, data_len):
        if i % 2 == 0:
            style_cmds.append(('BACKGROUND', (0, i), (-1, i), _CORP_ROW_ALT))
    return TableStyle(style_cmds)

@csrf_exempt
@require_http_methods(["POST"])
def update_inventory(request, inventory_name='default'):
    """
    Endpoint HTTP para importación de inventario.

    La lógica de negocio y procesamiento de archivos se delega al
    servicio `inventory.services.import_service`.
    """
    return procesar_importacion_inventario(request, inventory_name)

@require_http_methods(["GET"])
def get_monthly_movements(request):
    """
    Retorna entradas, salidas y saldo de cierre de los últimos 12 meses.
    """
    inventory_name = request.GET.get('inventory_name', 'default')
    warehouse_filter = request.GET.get('warehouse', '')
    category_filter = request.GET.get('category', '')
    search_filter = request.GET.get('search', '')

    try:
        result_data = get_monthly_movements_data(
            inventory_name=inventory_name,
            warehouse_filter=warehouse_filter,
            category_filter=category_filter,
            search_filter=search_filter,
        )
        return JsonResponse(result_data, safe=False)
    except Exception as e:
        logger.error(f"Error in get_monthly_movements: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["GET"])
def get_monthly_cuts(request):
    """
    Retorna cortes mensuales de inventario con promedio por mes.
    """
    inventory_name = request.GET.get('inventory_name', 'default')
    warehouse_filter = request.GET.get('warehouse', '')
    category_filter = request.GET.get('category', '')
    search_filter = request.GET.get('search', '')
    months = request.GET.get('months', '12')

    try:
        result_data = get_monthly_cuts_data(
            inventory_name=inventory_name,
            warehouse_filter=warehouse_filter,
            category_filter=category_filter,
            search_filter=search_filter,
            months=months,
        )
        return JsonResponse(result_data)
    except Exception as e:
        logger.error(f"Error in get_monthly_cuts: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["GET"])
def get_monthly_product_cuts(request):
    """
    Retorna corte mensual por producto (inventario promedio del mes).
    """
    inventory_name = request.GET.get('inventory_name', 'default')
    warehouse_filter = request.GET.get('warehouse', '')
    category_filter = request.GET.get('category', '')
    search_filter = request.GET.get('search', '')
    month = request.GET.get('month', '')
    limit = request.GET.get('limit', '')

    try:
        result_data = get_monthly_product_cuts_data(
            inventory_name=inventory_name,
            target_month=month,
            warehouse_filter=warehouse_filter,
            category_filter=category_filter,
            search_filter=search_filter,
            limit=limit,
        )
        return JsonResponse(result_data)
    except Exception as e:
        logger.error(f"Error in get_monthly_product_cuts: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["GET"])
def get_product_analysis(request):
    """
    Retorna el análisis de productos con filtros de rotación y estancamiento.
    """
    inventory_name = request.GET.get('inventory_name', 'default')
    category_filter = request.GET.get('category', '')
    warehouse_filter = request.GET.get('warehouse', '')
    rotation_filter = request.GET.get('rotation', '')
    stagnant_filter = request.GET.get('stagnant', '')
    high_rotation_filter = request.GET.get('high_rotation', '')
    exact_date = request.GET.get('date', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    search_filter = request.GET.get('search', '')
    limit = request.GET.get('limit', '')

    if exact_date and not date_to:
        # Filtro de día exacto: análisis como corte al cierre de esa fecha.
        date_to = exact_date

    try:
        analysis_data = get_product_analysis_data(
            inventory_name=inventory_name,
            category_filter=category_filter,
            warehouse_filter=warehouse_filter,
            rotation_filter=rotation_filter,
            stagnant_filter=stagnant_filter,
            high_rotation_filter=high_rotation_filter,
            date_from=date_from,
            date_to=date_to,
            search_filter=search_filter,
            limit=limit,
        )
        return JsonResponse(analysis_data, safe=False)
    except Exception as e:
        logger.error(f"Error in product analysis: {str(e)}", exc_info=True)
        return JsonResponse([], safe=False)



@require_http_methods(["GET"])
def get_batches(request):
    """
    Endpoint para obtener información de las importaciones realizadas.
    """
    inventory_name = request.GET.get('inventory_name', 'default')
    batches = ImportBatch.objects.filter(inventory_name=inventory_name).order_by('-started_at')
    
    batches_data = [{
        'id': batch.id,
        'file_name': batch.file_name,
        'inventory_name': batch.inventory_name,
        'started_at': batch.started_at.isoformat(),
        'processed_at': batch.processed_at.isoformat() if batch.processed_at else None,
        'rows_imported': batch.rows_imported,
        'rows_total': batch.rows_total,
        'checksum': batch.checksum,
    } for batch in batches]
    return JsonResponse(batches_data, safe=False)

@require_http_methods(["GET"])
def get_products(request):
    """
    Endpoint para obtener información de los productos.
    """
    inventory_name = request.GET.get('inventory_name', 'default')
    products = Product.objects.filter(inventory_name=inventory_name)
    products_data = [{
        'code': p.code,
        'description': p.description,
        'group': p.group,
        'initial_balance': float(p.initial_balance),
        'initial_unit_cost': float(p.initial_unit_cost),
    } for p in products]
    return JsonResponse(products_data, safe=False)

@require_http_methods(["GET"])
def get_records(request):
    """
    Endpoint para obtener los movimientos de inventario con filtros y paginación.
    """
    inventory_name = request.GET.get('inventory_name', 'default')
    warehouse_filter = request.GET.get('warehouse', '')
    category_filter = request.GET.get('category', '')
    exact_date = request.GET.get('date', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    search_filter = request.GET.get('search', '')

    try:
        records_query = InventoryRecord.objects.filter(
            product__inventory_name=inventory_name
        ).select_related('product', 'batch').only(
            'id', 'warehouse', 'date', 'document_type', 'document_number',
            'quantity', 'unit_cost', 'total', 'category',
            'product__code', 'product__description', 'batch__id'
        )

        # Aplicar filtros
        if warehouse_filter:
            records_query = records_query.filter(warehouse__icontains=warehouse_filter)
        if category_filter:
            records_query = records_query.filter(category__icontains=category_filter)
        if exact_date:
            records_query = records_query.filter(date=exact_date)
        else:
            if date_from:
                records_query = records_query.filter(date__gte=date_from)
            if date_to:
                records_query = records_query.filter(date__lte=date_to)
        if search_filter:
            records_query = records_query.filter(
                Q(product__code__icontains=search_filter) | Q(product__description__icontains=search_filter)
            )

        # Aplicar paginación
        records = records_query.order_by('-date')[:1000]
        records_data = [{
            'id': r.id,
            'product_code': r.product.code,
            'product_description': r.product.description,
            'warehouse': r.warehouse,
            'date': r.date.isoformat(),
            'document_type': r.document_type,
            'document_number': r.document_number,
            'quantity': float(r.quantity),
            'unit_cost': float(r.unit_cost),
            'total': float(r.total),
            'category': r.category,
            'batch_id': r.batch.id,
        } for r in records]
        return JsonResponse(records_data, safe=False)

    except Exception as e:
        logger.error(f"Error en obtener registros: {str(e)}", exc_info=True)
        return JsonResponse([], safe=False)


@require_http_methods(["POST"])
def create_inventory(request):
    """
    Endpoint para crear un nuevo inventario. Verifica que el nombre no exista
    """
    try:
        data = json.loads(request.body)
        inventory_name = data.get('inventory_name', '').strip().lower()
        if not inventory_name:
            return JsonResponse({'ok': False, 'error': 'Nombre de inventario requerido'}, status=400)

        # Verifica si ya existe un inventario con el mismo nombre que uno existente en la base de datos
        if Product.objects.filter(inventory_name=inventory_name).exists():
            return JsonResponse({'ok': False, 'error': 'El inventario ya existe'}, status=400)

        return JsonResponse({'ok': True, 'inventory_name': inventory_name})
    except Exception as e:
        logger.error(f"Error creating inventory: {str(e)}", exc_info=True)
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)


@require_http_methods(["GET"])
def get_product_history(request, product_code, inventory_name='default'):
    """
    Endpoint para obtener el historial de un producto.
    """
    try:
        records = InventoryRecord.objects.filter(
            product__code=product_code,
            product__inventory_name=inventory_name
        ).select_related('product', 'batch').order_by('date')

        history_data = [{
            'id': r.id,
            'date': r.date.isoformat(),
            'quantity': float(r.quantity),
            'unit_cost': float(r.unit_cost),
            'total': float(r.total),
            'warehouse': r.warehouse,
            'document_type': r.document_type,
            'document_number': r.document_number,
            'batch_id': r.batch.id,
        } for r in records]

        return JsonResponse(history_data, safe=False)
    except Exception as e:
        logger.error(f"Error retrieving product history: {str(e)}", exc_info=True)
        return JsonResponse([], safe=False)


@require_http_methods(["GET"])
def get_summary(request):
    """
    Endpoint para obtener el resumen del inventario, incluyendo métricas clave como:
    - Total de productos
    - Total de inventario
    """
    inventory_name = request.GET.get('inventory_name', 'default')
    try:
        summary_data = get_inventory_summary_data(inventory_name=inventory_name)
        return JsonResponse(summary_data)
    except Exception as e:
        logger.error(f"Error retrieving summary: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["GET"])
def export_analysis(request, inventory_name='default'):
    """
    Endpoint para exportar el análisis de productos.
    """
    try:
        # Obtener el tipo de formato deseado
        format_type = request.GET.get('format', 'excel')

        #
        analysis_response = get_product_analysis(request)
        analysis_data = analysis_response.content.decode('utf-8')
        analysis_list = json.loads(analysis_data)

        logo_path = _get_logo_path()

        if format_type == 'excel':
            wb = Workbook()
            ws = wb.active
            ws.title = 'Analisis'

            data_row = _apply_excel_header(ws, 'Análisis de Inventario', inventory_name, logo_path)

            col_headers = [
                'Código', 'Producto', 'Grupo', 'Cantidad Actual', 'Valor Actual',
                'Costo Unitario', 'Consumido', 'Estancado', 'Rotación', 'Alta Rotación', 'Almacén',
            ]
            col_widths_xl = [15, 40, 20, 18, 18, 18, 12, 12, 15, 15, 25]
            for ci, (hdr, w) in enumerate(zip(col_headers, col_widths_xl), 1):
                cell = ws.cell(row=data_row, column=ci, value=hdr)
                ws.column_dimensions[get_column_letter(ci)].width = w

            for ri, item in enumerate(analysis_list, 1):
                rn = data_row + ri
                ws.cell(row=rn, column=1,  value=str(item.get('codigo', '')))
                ws.cell(row=rn, column=2,  value=str(item.get('nombre_producto', '')))
                ws.cell(row=rn, column=3,  value=str(item.get('grupo', '')))
                ws.cell(row=rn, column=4,  value=float(item.get('cantidad_saldo_actual', 0)))
                ws.cell(row=rn, column=5,  value=float(item.get('valor_saldo_actual', 0)))
                ws.cell(row=rn, column=6,  value=float(item.get('costo_unitario', 0)))
                ws.cell(row=rn, column=7,  value=str(item.get('consumed', '')))
                ws.cell(row=rn, column=8,  value=str(item.get('estancado', '')))
                ws.cell(row=rn, column=9,  value=str(item.get('rotacion', '')))
                ws.cell(row=rn, column=10, value=str(item.get('alta_rotacion', '')))
                ws.cell(row=rn, column=11, value=str(item.get('almacen', '')))

            data_end = data_row + len(analysis_list)
            _style_excel_table(ws, data_row, data_row + 1, data_end, len(col_headers))

            # Hoja de gráfica Pie: distribución de rotación
            chart_ws = wb.create_sheet('GraficaAnalisis')
            estancados_c    = sum(1 for x in analysis_list if x.get('estancado'))
            alta_rot_c      = sum(1 for x in analysis_list if x.get('alta_rotacion'))
            normal_c        = len(analysis_list) - estancados_c - alta_rot_c

            chart_ws['A1'] = 'Distribución de Rotación de Inventario'
            chart_ws['A1'].font = Font(name='Calibri', size=13, bold=True, color=_XL_ACCENT)
            chart_ws['A3'] = 'Categoría'
            chart_ws['B3'] = 'Cantidad'
            chart_ws['A4'] = 'Estancados'
            chart_ws['B4'] = estancados_c
            chart_ws['A5'] = 'Alta Rotación'
            chart_ws['B5'] = alta_rot_c
            chart_ws['A6'] = 'Normal'
            chart_ws['B6'] = normal_c

            if len(analysis_list) > 0:
                pie = PieChart()
                pie.title  = 'Distribución de Rotación'
                pie.style  = 10
                pie.width  = 18
                pie.height = 12
                labels_ref = Reference(chart_ws, min_col=1, min_row=4, max_row=6)
                data_ref   = Reference(chart_ws, min_col=2, min_row=3, max_row=6)
                pie.add_data(data_ref, titles_from_data=True)
                pie.set_categories(labels_ref)
                slice_colors = ['EF7C91', 'EF4444', 'EBF9FC']
                for si, color_hex in enumerate(slice_colors):
                    dp = XLDataPoint(idx=si)
                    dp.graphicalProperties.solidFill = color_hex
                    pie.series[0].dPt.append(dp)
                chart_ws.add_chart(pie, 'D3')

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="inventory_analysis_{inventory_name}.xlsx"'
            return response

        elif format_type == 'pdf':
            buffer = BytesIO()
            doc = _create_pdf_doc(buffer, landscape(A4), 'Análisis de Inventario')
            elements = _build_pdf_header_elements(logo_path, 'Análisis de Inventario', inventory_name)
            elements.append(Spacer(1, 8))

            rl_styles = getSampleStyleSheet()
            normal_style = ParagraphStyle(
                'AnaNormal', parent=rl_styles['Normal'],
                fontName='Times-Roman', fontSize=7,
                wordWrap='LTR', splitLongWords=True, leading=9,
            )
            header_style = ParagraphStyle(
                'AnaHeader', parent=rl_styles['Normal'],
                fontName='Times-Bold', fontSize=9, alignment=1,
            )

            # Gráfica Pie de distribución de rotación
            if analysis_list:
                estancados_c = sum(1 for x in analysis_list if x.get('estancado'))
                alta_rot_c   = sum(1 for x in analysis_list if x.get('alta_rotacion'))
                normal_c     = len(analysis_list) - estancados_c - alta_rot_c

                pie_drawing = Drawing(380, 170)
                rl_pie = RLPie()
                rl_pie.x      = 50
                rl_pie.y      = 25
                rl_pie.width  = 120
                rl_pie.height = 120
                rl_pie.data   = [max(estancados_c, 0), max(alta_rot_c, 0), max(normal_c, 0)]
                rl_pie.labels = [
                    f'Estancados ({estancados_c})',
                    f'Alta Rot. ({alta_rot_c})',
                    f'Normal ({normal_c})',
                ]
                rl_pie.slices[0].fillColor = rl_colors.HexColor('#EF4444')
                rl_pie.slices[1].fillColor = rl_colors.HexColor('#EF7C91')
                rl_pie.slices[2].fillColor = rl_colors.HexColor('#4BC0D9')
                rl_pie.slices.strokeColor  = rl_colors.white
                rl_pie.slices.strokeWidth  = 0.5
                legend = Legend()
                legend.x = 200
                legend.y = 80
                legend.dx = 10
                legend.dy = 10
                legend.fontName  = 'Times-Roman'
                legend.fontSize  = 8
                legend.colorNamePairs = [
                    (rl_colors.HexColor('#EF4444'), f'Estancados ({estancados_c})'),
                    (rl_colors.HexColor('#EF7C91'), f'Alta Rotación ({alta_rot_c})'),
                    (rl_colors.HexColor('#4BC0D9'), f'Normal ({normal_c})'),
                ]
                pie_drawing.add(rl_pie)
                pie_drawing.add(legend)
                elements.append(GraphicsFlowable(pie_drawing))
                elements.append(Spacer(1, 10))

            # Tabla de datos
            if analysis_list:
                headers = [
                    Paragraph('Código', header_style),
                    Paragraph('Producto', header_style),
                    Paragraph('Grupo', header_style),
                    Paragraph('Cantidad', header_style),
                    Paragraph('Valor', header_style),
                    Paragraph('Costo Unit.', header_style),
                    Paragraph('Estancado', header_style),
                    Paragraph('Rotación', header_style),
                    Paragraph('Alta Rot.', header_style),
                    Paragraph('Almacén', header_style),
                ]
                rows = []
                for item in analysis_list:
                    rows.append([
                        Paragraph(str(item.get('codigo', '')), normal_style),
                        Paragraph(str(item.get('nombre_producto', '')), normal_style),
                        Paragraph(str(item.get('grupo', '')), normal_style),
                        Paragraph(f"{item.get('cantidad_saldo_actual', 0):,.2f}", normal_style),
                        Paragraph(f"${item.get('valor_saldo_actual', 0):,.2f}", normal_style),
                        Paragraph(f"${item.get('costo_unitario', 0):,.2f}", normal_style),
                        Paragraph(str(item.get('estancado', '')), normal_style),
                        Paragraph(str(item.get('rotacion', '')), normal_style),
                        Paragraph(str(item.get('alta_rotacion', '')), normal_style),
                        Paragraph(str(item.get('almacen', '')), normal_style),
                    ])
                data = [headers] + rows
                colWidths = [57, 130, 71, 68, 78, 71, 57, 55, 55, 83]
                table = Table(data, colWidths=colWidths, repeatRows=1)
                table.setStyle(_pdf_table_style(len(data)))
                elements.append(table)

            try:
                doc.build(elements)
                buffer.seek(0)
                if not buffer.getvalue():
                    logger.error("PDF buffer vacío tras build")
                    return JsonResponse({'error': 'PDF generation failed - empty content'}, status=500)
            except Exception as e:
                logger.error(f"Error building PDF analysis: {str(e)}", exc_info=True)
                return JsonResponse({'error': f'PDF generation failed: {str(e)}'}, status=500)

            response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="inventory_analysis_{inventory_name}.pdf"'
            return response
        else:
            return JsonResponse({'error': 'Formato no soportado'}, status=400)
    except Exception as e:
        logger.error(f"Error exporting analysis: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["GET"])
def export_movements(request, inventory_name='default'):
    """
    Exportar movimientos
    """
    try:
        # query para obtener formato de el excel
        format_type = request.GET.get('format', 'excel')

        # query para consultar datos de movimientos por filtros
        inventory_name_param = request.GET.get('inventory_name', inventory_name)
        warehouse_filter = request.GET.get('warehouse', '')
        category_filter = request.GET.get('category', '')
        exact_date = request.GET.get('date', '')
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        search_filter = request.GET.get('search', '')

        records_query = InventoryRecord.objects.filter(product__inventory_name=inventory_name_param).select_related('product', 'batch')

        # Aplicacion de filtros
        if warehouse_filter:
            records_query = records_query.filter(warehouse__icontains=warehouse_filter)
        if category_filter:
            records_query = records_query.filter(category__icontains=category_filter)
        if exact_date:
            records_query = records_query.filter(date=exact_date)
        else:
            if date_from:
                records_query = records_query.filter(date__gte=date_from)
            if date_to:
                records_query = records_query.filter(date__lte=date_to)
        if search_filter:
            records_query = records_query.filter(
                Q(product__code__icontains=search_filter) | Q(product__description__icontains=search_filter)
            )

        # Limite de exportacion de hasta  5000 de record(Historial)
        records = records_query.order_by('-date')[:5000]
        movements_data = [{
            'fecha': r.date.isoformat(),
            'codigo': r.product.code,
            'nombre_producto': r.product.description,
            'almacen': r.warehouse,
            'tipo_documento': r.document_type,
            'documento': r.document_number,
            'cantidad': float(r.quantity),
            'costo_unitario': float(r.unit_cost),
            'costo_total': float(r.total),
            'categoria': r.category,
        } for r in records]

        logo_path = _get_logo_path()

        if format_type == 'excel':
            wb = Workbook()
            ws = wb.active
            ws.title = 'Movimientos'

            data_row = _apply_excel_header(ws, 'Movimientos de Inventario', inventory_name_param, logo_path)

            col_headers = ['Fecha', 'Código', 'Producto', 'Almacén', 'Tipo Doc.', 'Documento', 'Cantidad', 'Costo Unit.', 'Total', 'Categoría']
            col_widths_xl = [12, 14, 38, 20, 14, 16, 14, 16, 16, 18]
            for ci, (hdr, w) in enumerate(zip(col_headers, col_widths_xl), 1):
                ws.cell(row=data_row, column=ci, value=hdr)
                ws.column_dimensions[get_column_letter(ci)].width = w

            for ri, item in enumerate(movements_data, 1):
                rn = data_row + ri
                ws.cell(row=rn, column=1,  value=item['fecha'])
                ws.cell(row=rn, column=2,  value=item['codigo'])
                ws.cell(row=rn, column=3,  value=item['nombre_producto'])
                ws.cell(row=rn, column=4,  value=item['almacen'])
                ws.cell(row=rn, column=5,  value=item['tipo_documento'] or '')
                ws.cell(row=rn, column=6,  value=item['documento'] or '')
                ws.cell(row=rn, column=7,  value=item['cantidad'])
                ws.cell(row=rn, column=8,  value=item['costo_unitario'])
                ws.cell(row=rn, column=9,  value=item['costo_total'])
                ws.cell(row=rn, column=10, value=item['categoria'])

            data_end = data_row + len(movements_data)
            _style_excel_table(ws, data_row, data_row + 1, data_end, len(col_headers))

            # Hoja de gráfica: movimientos agrupados por mes
            monthly_totals = defaultdict(lambda: {'entradas': 0.0, 'salidas': 0.0})
            for item in movements_data:
                mk = str(item['fecha'])[:7]
                if item['cantidad'] >= 0:
                    monthly_totals[mk]['entradas'] += item['cantidad']
                else:
                    monthly_totals[mk]['salidas'] += abs(item['cantidad'])

            chart_ws = wb.create_sheet('GraficaMovimientos')
            months_sorted = sorted(monthly_totals.keys())
            chart_ws['A1'] = 'Movimientos por Mes'
            chart_ws['A1'].font = Font(name='Calibri', size=13, bold=True, color=_XL_ACCENT)
            chart_ws['A3'] = 'Mes'
            chart_ws['B3'] = 'Entradas'
            chart_ws['C3'] = 'Salidas'
            for i, mk in enumerate(months_sorted, 4):
                chart_ws.cell(row=i, column=1, value=mk)
                chart_ws.cell(row=i, column=2, value=monthly_totals[mk]['entradas'])
                chart_ws.cell(row=i, column=3, value=monthly_totals[mk]['salidas'])

            if months_sorted:
                bar = BarChart()
                bar.type      = 'col'
                bar.grouping  = 'clustered'
                bar.title     = 'Movimientos por Mes'
                bar.style     = 10
                bar.y_axis.title = 'Cantidad'
                bar.x_axis.title = 'Mes'
                bar.width  = 22
                bar.height = 14
                data_ref = Reference(chart_ws, min_col=2, max_col=3, min_row=3, max_row=3 + len(months_sorted))
                cats_ref = Reference(chart_ws, min_col=1, min_row=4, max_row=3 + len(months_sorted))
                bar.add_data(data_ref, titles_from_data=True)
                bar.set_categories(cats_ref)
                bar.series[0].graphicalProperties.solidFill = 'EF7C91'
                bar.series[1].graphicalProperties.solidFill = '4BC0D9'
                chart_ws.add_chart(bar, 'E3')

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="movimientos_inventario_{inventory_name_param}.xlsx"'
            return response

        elif format_type == 'pdf':
            buffer = BytesIO()
            doc = _create_pdf_doc(buffer, landscape(A4), 'Movimientos de Inventario')
            elements = _build_pdf_header_elements(logo_path, 'Movimientos de Inventario', inventory_name_param)
            elements.append(Spacer(1, 8))

            rl_styles = getSampleStyleSheet()
            normal_style = ParagraphStyle(
                'MovNormal', parent=rl_styles['Normal'],
                fontName='Times-Roman', fontSize=7,
                wordWrap='LTR', splitLongWords=True, leading=9,
            )
            header_style = ParagraphStyle(
                'MovHeader', parent=rl_styles['Normal'],
                fontName='Times-Bold', fontSize=9, alignment=1,
            )

            # Gráfica VerticalBarChart agrupada por mes
            monthly_totals = defaultdict(lambda: {'entradas': 0.0, 'salidas': 0.0})
            for item in movements_data:
                mk = str(item['fecha'])[:7]
                if item['cantidad'] >= 0:
                    monthly_totals[mk]['entradas'] += item['cantidad']
                else:
                    monthly_totals[mk]['salidas'] += abs(item['cantidad'])
            months_sorted = sorted(monthly_totals.keys())[-12:]
            entradas_vals = [monthly_totals[m]['entradas'] for m in months_sorted]
            salidas_vals  = [monthly_totals[m]['salidas']  for m in months_sorted]

            if months_sorted and (any(entradas_vals) or any(salidas_vals)):
                bar_drawing = Drawing(620, 200)
                bc = VerticalBarChart()
                bc.x      = 60
                bc.y      = 20
                bc.height = 160
                bc.width  = 540
                bc.data   = [entradas_vals, salidas_vals]
                bc.categoryAxis.categoryNames = [m[-5:] for m in months_sorted]
                bc.bars[0].fillColor = rl_colors.HexColor('#EF7C91')
                bc.bars[1].fillColor = rl_colors.HexColor('#4BC0D9')
                bc.valueAxis.valueMin = 0
                bar_drawing.add(bc)
                elements.append(GraphicsFlowable(bar_drawing))
                elements.append(Spacer(1, 10))

            # Tabla de movimientos
            if movements_data:
                headers = [
                    Paragraph('Fecha', header_style),
                    Paragraph('Código', header_style),
                    Paragraph('Producto', header_style),
                    Paragraph('Almacén', header_style),
                    Paragraph('Tipo Doc.', header_style),
                    Paragraph('Documento', header_style),
                    Paragraph('Cantidad', header_style),
                    Paragraph('Costo Unit.', header_style),
                    Paragraph('Total', header_style),
                    Paragraph('Categoría', header_style),
                ]
                rows = []
                for item in movements_data:
                    rows.append([
                        Paragraph(str(item['fecha']), normal_style),
                        Paragraph(str(item['codigo']), normal_style),
                        Paragraph(str(item['nombre_producto']), normal_style),
                        Paragraph(str(item['almacen']), normal_style),
                        Paragraph(str(item['tipo_documento'] or ''), normal_style),
                        Paragraph(str(item['documento'] or ''), normal_style),
                        Paragraph(f"{item['cantidad']:,.2f}", normal_style),
                        Paragraph(f"${item['costo_unitario']:,.2f}", normal_style),
                        Paragraph(f"${item['costo_total']:,.2f}", normal_style),
                        Paragraph(str(item['categoria']), normal_style),
                    ])
                data = [headers] + rows
                colWidths = [68, 58, 140, 76, 58, 58, 72, 82, 82, 66]
                table = Table(data, colWidths=colWidths, repeatRows=1)
                table.setStyle(_pdf_table_style(len(data)))
                elements.append(table)

            try:
                doc.build(elements)
                buffer.seek(0)
                if not buffer.getvalue():
                    logger.error("PDF buffer vacío tras build")
                    return JsonResponse({'error': 'PDF generation failed - empty content'}, status=500)
            except Exception as e:
                logger.error(f"Error building PDF movements: {str(e)}", exc_info=True)
                return JsonResponse({'error': f'PDF generation failed: {str(e)}'}, status=500)

            response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="movimientos_inventario_{inventory_name_param}.pdf"'
            return response
        else:
            return JsonResponse({'error': 'Formato no soportado'}, status=400)
    except Exception as e:
        logger.error(f"Error exporting movements: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["GET"])
def export_monthly_cuts(request, inventory_name='default'):
    """
    Exporta el corte mensual promedio en Excel o PDF.
    """
    try:
        format_type = request.GET.get('format', 'excel')
        inventory_name_param = request.GET.get('inventory_name', inventory_name)
        warehouse_filter = request.GET.get('warehouse', '')
        category_filter = request.GET.get('category', '')
        search_filter = request.GET.get('search', '')
        months = request.GET.get('months', '12')
        month = request.GET.get('month', '')
        product_limit = request.GET.get('product_limit', '5000')

        cuts_payload = get_monthly_cuts_data(
            inventory_name=inventory_name_param,
            warehouse_filter=warehouse_filter,
            category_filter=category_filter,
            search_filter=search_filter,
            months=months,
        )
        cuts_rows = cuts_payload.get('months', [])
        period_average_general = cuts_payload.get('period_average_general', 0)
        product_cuts_payload = get_monthly_product_cuts_data(
            inventory_name=inventory_name_param,
            target_month=month,
            warehouse_filter=warehouse_filter,
            category_filter=category_filter,
            search_filter=search_filter,
            limit=product_limit,
        )
        product_rows = product_cuts_payload.get('products', [])

        export_rows = []
        for row in cuts_rows:
            export_rows.append({
                'mes': row.get('month'),
                'corte_inicial': row.get('opening_balance', 0),
                'entradas': row.get('total_entries', 0),
                'salidas': row.get('total_exits', 0),
                'corte_final': row.get('closing_balance', 0),
                'corte_promedio_general': row.get('average_balance_general', row.get('average_balance', 0)),
            })

        logo_path = _get_logo_path()

        if format_type == 'excel':
            wb = Workbook()

            # --- Hoja CortesMensuales ---
            ws = wb.active
            ws.title = 'CortesMensuales'
            data_row = _apply_excel_header(ws, 'Cortes Mensuales de Inventario', inventory_name_param, logo_path)

            cuts_col_headers = ['Mes', 'Corte Inicial', 'Entradas', 'Salidas', 'Corte Final', 'Promedio General']
            cuts_col_widths  = [14, 18, 16, 16, 18, 22]
            for ci, (hdr, w) in enumerate(zip(cuts_col_headers, cuts_col_widths), 1):
                ws.cell(row=data_row, column=ci, value=hdr)
                ws.column_dimensions[get_column_letter(ci)].width = w

            for ri, row in enumerate(export_rows, 1):
                rn = data_row + ri
                ws.cell(row=rn, column=1, value=str(row['mes']))
                ws.cell(row=rn, column=2, value=float(row['corte_inicial']))
                ws.cell(row=rn, column=3, value=float(row['entradas']))
                ws.cell(row=rn, column=4, value=float(row['salidas']))
                ws.cell(row=rn, column=5, value=float(row['corte_final']))
                ws.cell(row=rn, column=6, value=float(row['corte_promedio_general']))

            cuts_end = data_row + len(export_rows)
            _style_excel_table(ws, data_row, data_row + 1, cuts_end, len(cuts_col_headers))

            # Promedios del periodo (celdas extra)
            ws.column_dimensions['I'].width = 30
            ws.column_dimensions['J'].width = 22
            ws['I1'] = 'Promedio del periodo (general)'
            ws['I1'].font = Font(name='Calibri', size=10, bold=True, color=_XL_ACCENT)
            ws['J1'] = float(period_average_general)

            # Hoja de gráfica LineChart
            chart_ws = wb.create_sheet('GraficaCortes')
            chart_ws['A1'] = 'Evolución de Cortes Mensuales'
            chart_ws['A1'].font = Font(name='Calibri', size=13, bold=True, color=_XL_ACCENT)
            chart_ws['A3'] = 'Mes'
            chart_ws['B3'] = 'Corte Inicial'
            chart_ws['C3'] = 'Corte Final'
            chart_ws['D3'] = 'Promedio General'
            for i, row in enumerate(export_rows, 4):
                chart_ws.cell(row=i, column=1, value=str(row['mes']))
                chart_ws.cell(row=i, column=2, value=float(row['corte_inicial']))
                chart_ws.cell(row=i, column=3, value=float(row['corte_final']))
                chart_ws.cell(row=i, column=4, value=float(row['corte_promedio_general']))

            if export_rows:
                line = LineChart()
                line.title      = 'Evolución de Cortes Mensuales'
                line.style      = 10
                line.y_axis.title = 'Valor (COP)'
                line.x_axis.title = 'Mes'
                line.width  = 22
                line.height = 14
                data_ref = Reference(chart_ws, min_col=2, max_col=4, min_row=3, max_row=3 + len(export_rows))
                cats_ref = Reference(chart_ws, min_col=1, min_row=4, max_row=3 + len(export_rows))
                line.add_data(data_ref, titles_from_data=True)
                line.set_categories(cats_ref)
                line.series[0].graphicalProperties.line.solidFill = 'EF7C91'
                line.series[1].graphicalProperties.line.solidFill = '4BC0D9'
                line.series[2].graphicalProperties.line.solidFill = 'D9607A'
                chart_ws.add_chart(line, 'F3')

            # --- Hoja CorteProductosMes ---
            if product_rows:
                prod_ws = wb.create_sheet('CorteProductosMes')
                prod_data_row = _apply_excel_header(prod_ws, 'Corte de Productos por Mes', inventory_name_param, logo_path)

                prod_keys = list(product_rows[0].keys()) if product_rows else []
                prod_col_widths_map = {'A': 14, 'B': 38, 'C': 24, 'D': 16, 'E': 16, 'F': 16, 'G': 14, 'H': 18, 'I': 18, 'J': 18}
                for ci, key in enumerate(prod_keys, 1):
                    prod_ws.cell(row=prod_data_row, column=ci, value=key)
                    col_letter = get_column_letter(ci)
                    prod_ws.column_dimensions[col_letter].width = prod_col_widths_map.get(col_letter, 14)

                for ri, prow in enumerate(product_rows, 1):
                    rn = prod_data_row + ri
                    for ci, key in enumerate(prod_keys, 1):
                        prod_ws.cell(row=rn, column=ci, value=prow.get(key, ''))

                prod_end = prod_data_row + len(product_rows)
                _style_excel_table(prod_ws, prod_data_row, prod_data_row + 1, prod_end, len(prod_keys))

                totals = product_cuts_payload.get('totals', {})
                prod_ws.column_dimensions['L'].width = 18
                prod_ws.column_dimensions['M'].width = 18
                prod_ws['L1'] = 'Mes del corte'
                prod_ws['L1'].font = Font(name='Calibri', size=10, bold=True, color=_XL_ACCENT)
                prod_ws['M1'] = product_cuts_payload.get('month', '')
                prod_ws['L3'] = 'Totales valor'
                prod_ws['L3'].font = Font(name='Calibri', size=10, bold=True, color=_XL_ACCENT)
                prod_ws['L4'] = 'Apertura'
                prod_ws['M4'] = float(totals.get('opening_value', 0))
                prod_ws['L5'] = 'Cierre'
                prod_ws['M5'] = float(totals.get('closing_value', 0))
                prod_ws['L6'] = 'Promedio'
                prod_ws['M6'] = float(totals.get('average_value', 0))

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = (
                f'attachment; filename="cortes_mensuales_{inventory_name_param}.xlsx"'
            )
            return response

        if format_type == 'pdf':
            buffer = BytesIO()
            doc = _create_pdf_doc(buffer, landscape(A4), 'Cortes Mensuales')
            rl_styles = getSampleStyleSheet()
            normal_style = ParagraphStyle(
                'CutsNormal', parent=rl_styles['Normal'],
                fontName='Times-Roman', fontSize=8, leading=10,
            )
            header_style = ParagraphStyle(
                'CutsHeader', parent=rl_styles['Normal'],
                fontName='Times-Bold', fontSize=9, alignment=1,
            )
            info_style = ParagraphStyle(
                'CutsInfo', parent=rl_styles['Normal'],
                fontName='Times-Italic', fontSize=9,
                textColor=rl_colors.HexColor('#3D4459'),
            )

            elements = _build_pdf_header_elements(logo_path, 'Cortes Mensuales de Inventario', inventory_name_param)
            elements += [
                Spacer(1, 6),
                Paragraph(f'Promedio del periodo (general): ${float(period_average_general):,.2f}', info_style),
                Spacer(1, 8),
            ]

            # Gráfica de línea con evolución del corte
            if export_rows:
                closing  = [float(r['corte_final'])         for r in export_rows]
                avg_gen  = [float(r['corte_promedio_general']) for r in export_rows]
                if any(closing) or any(avg_gen):
                    line_drawing = Drawing(620, 190)
                    lc = HorizontalLineChart()
                    lc.x      = 60
                    lc.y      = 20
                    lc.height = 150
                    lc.width  = 540
                    lc.data   = [closing, avg_gen]
                    lc.categoryAxis.categoryNames = [str(r['mes'])[-7:] for r in export_rows]
                    lc.lines[0].strokeColor = rl_colors.HexColor('#EF7C91')
                    lc.lines[0].strokeWidth = 1.5
                    lc.lines[1].strokeColor = rl_colors.HexColor('#4BC0D9')
                    lc.lines[1].strokeWidth = 1.5
                    lc.valueAxis.valueMin = 0
                    line_drawing.add(lc)
                    elements.append(GraphicsFlowable(line_drawing))
                    elements.append(Spacer(1, 10))

            # Tabla de cortes mensuales
            cuts_headers = [
                Paragraph('Mes', header_style),
                Paragraph('Corte Inicial', header_style),
                Paragraph('Entradas', header_style),
                Paragraph('Salidas', header_style),
                Paragraph('Corte Final', header_style),
                Paragraph('Prom. General', header_style),
            ]
            table_rows_pdf = [cuts_headers]
            for row in export_rows:
                table_rows_pdf.append([
                    Paragraph(str(row['mes']), normal_style),
                    Paragraph(f"${float(row['corte_inicial']):,.2f}", normal_style),
                    Paragraph(f"${float(row['entradas']):,.2f}", normal_style),
                    Paragraph(f"${float(row['salidas']):,.2f}", normal_style),
                    Paragraph(f"${float(row['corte_final']):,.2f}", normal_style),
                    Paragraph(f"${float(row['corte_promedio_general']):,.2f}", normal_style),
                ])

            table = Table(table_rows_pdf, colWidths=[75, 95, 85, 85, 95, 95], repeatRows=1)
            table.setStyle(_pdf_table_style(len(table_rows_pdf)))
            elements.append(table)

            try:
                doc.build(elements)
                buffer.seek(0)
                if not buffer.getvalue():
                    logger.error("PDF buffer vacío tras build")
                    return JsonResponse({'error': 'PDF generation failed - empty content'}, status=500)
            except Exception as e:
                logger.error(f"Error building PDF cuts: {str(e)}", exc_info=True)
                return JsonResponse({'error': f'PDF generation failed: {str(e)}'}, status=500)

            response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = (
                f'attachment; filename="cortes_mensuales_{inventory_name_param}.pdf"'
            )
            return response

        return JsonResponse({'error': 'Formato no soportado'}, status=400)

    except Exception as e:
        logger.error(f"Error exporting monthly cuts: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["GET"])
def list_inventories(request):
    """
    Lista todos los inventarios
    """
    try:
        inventories = Product.objects.values_list('inventory_name', flat=True).distinct()
        return JsonResponse(list(inventories), safe=False)
    except Exception as e:
        logger.error(f"Error listing inventories: {str(e)}", exc_info=True)
        return JsonResponse([], safe=False)


@csrf_exempt
@require_http_methods(["POST"])
def rollback_batch(request):
    """
    Revierte un lote de actualización eliminando sus movimientos importados.

    Nota:
    - Solo aplica a lotes que contengan movimientos.
    - No revierte automáticamente productos base (archivo inicial).
    """
    inventory_name = request.GET.get('inventory_name', 'default')

    try:
        body = json.loads(request.body) if request.body else {}
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'error': 'JSON inválido'}, status=400)

    inventory_name = str(body.get('inventory_name', inventory_name)).strip().lower() or 'default'
    batch_id = body.get('batch_id')

    try:
        batches_qs = ImportBatch.objects.filter(inventory_name=inventory_name).order_by('-processed_at', '-id')
        if batch_id:
            batches_qs = batches_qs.filter(id=batch_id)
        else:
            batch_ids_with_records = InventoryRecord.objects.filter(
                product__inventory_name=inventory_name
            ).values_list('batch_id', flat=True).distinct()
            batches_qs = batches_qs.filter(id__in=batch_ids_with_records)

        batch = batches_qs.first()
        if not batch:
            return JsonResponse(
                {'ok': False, 'error': 'No se encontró el lote a revertir para ese inventario'},
                status=404,
            )

        records_qs = InventoryRecord.objects.filter(
            batch=batch,
            product__inventory_name=inventory_name,
        )
        records_count = records_qs.count()
        if records_count == 0:
            return JsonResponse(
                {
                    'ok': False,
                    'error': 'El lote seleccionado no tiene movimientos para revertir (posible lote base).',
                },
                status=400,
            )

        deleted_rows, _ = records_qs.delete()
        batch_info = {
            'id': batch.id,
            'file_name': batch.file_name,
            'processed_at': batch.processed_at.isoformat() if batch.processed_at else None,
        }
        batch.delete()

        return JsonResponse(
            {
                'ok': True,
                'inventory_name': inventory_name,
                'rolled_back_batch': batch_info,
                'deleted_rows': deleted_rows,
                'deleted_records': records_count,
            }
        )
    except Exception as e:
        logger.error(f"Error rolling back batch: {str(e)}", exc_info=True)
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)



@require_http_methods(["GET"])
def get_last_update_time(request):
    """
    Recupera la marca de tiempo de la última actualización del inventario.
    """
    inventory_name = request.GET.get('inventory_name', 'default')
    try:
        last_batch = ImportBatch.objects.filter(
            inventory_name=inventory_name,
            processed_at__isnull=False
        ).order_by('-processed_at').first()

        if last_batch:
            return JsonResponse({'last_update': last_batch.processed_at.isoformat()})
        else:
            return JsonResponse({'last_update': None})
    except Exception as e:
        logger.error(f"Error retrieving last update time: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def upload_base_file(request, inventory_name='default'):
    """
    Actualiza la base de inventarios
    """
    # This is similar to update_inventory but only for base files
    return update_inventory(request, inventory_name)


@require_http_methods(["GET"])
def get_inventory_at_date(request):
    """
    Calcula el estado del inventario (cantidad y valor) para una fecha.
    """
    inventory_name = request.GET.get('inventory_name', 'default')
    date_str = request.GET.get('date', '')
    warehouse_filter = request.GET.get('warehouse', '')
    category_filter = request.GET.get('category', '')

    if not date_str:
        return JsonResponse({'error': 'Date parameter is required (format: YYYY-MM-DD)'}, status=400)

    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Invalid date format. Use YYYY-MM-DD'}, status=400)

    try:
        data = get_inventory_at_date_data(
            inventory_name=inventory_name,
            date_str=date_str,
            target_date=target_date,
            warehouse_filter=warehouse_filter,
            category_filter=category_filter,
        )
        return JsonResponse(data)

    except Exception as e:
        logger.error(f"Error calculating inventory at date: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["GET"])
def welcome(request):
    """
    Retorna bienvenido para pruebas de respuesta de la api
    """
    logger.info(f"Request received: {request.method} {request.path}")
    return JsonResponse({'message': 'Bienvenido a el sistema de analisis de inventarios'})
